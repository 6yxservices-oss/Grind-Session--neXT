"""
Deliverable 3: Content Intel Tracker Generator (.xlsx)
-------------------------------------------------------
Builds or appends to the 11-tab master content operations spreadsheet.
Critical rule: If the tracker already exists, APPEND with a black separator row.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill,
    Font,
    Alignment,
    Border,
    Side,
    NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Styles ────────────────────────────────────────────────────────────────────
NAVY_FILL = PatternFill(start_color="041E42", end_color="041E42", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
WHITE_BOLD = Font(name="Arial", bold=True, color="FFFFFF", size=11)
CENTER = Alignment(horizontal="center", vertical="center")
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ── Tab Definitions ───────────────────────────────────────────────────────────
TAB_DEFINITIONS: dict[str, dict] = {
    "Mike V — Twitter": {
        "color": "041E42",
        "columns": [
            "Drop Day", "Post Time", "Content Side", "Copy (Editable)",
            "Graphic/Asset Needed", "Graphic Status", "Asset Link (Google Drive)",
            "Link Included", "Sponsor Tie-In", "Post Status", "Actual Post Time",
            "Post URL", "Impressions", "Engagements", "Notes", "Screenshot / File",
        ],
    },
    "Mike V — Instagram": {
        "color": "041E42",
        "columns": [
            "Drop Day", "Post Time", "Format", "Content Side",
            "Caption / Copy (Editable)", "Graphic/Asset Needed", "Graphic Status",
            "Asset Link (Google Drive)", "Hashtags", "Link in Bio Updated?",
            "Sponsor Tie-In", "Status", "Actual Post Time", "Post URL",
            "Impressions", "Engagements", "Notes", "Screenshot / File",
        ],
    },
    "HVU Insider — Twitter": {
        "color": "041E42",
        "columns": [
            "Drop Day", "Post Time", "Post Type", "Content Side",
            "Copy (Editable)", "Graphic/Asset Needed", "Graphic Status",
            "Asset Link (Google Drive)", "Link Included", "Sponsor Tie-In",
            "Status", "Actual Post Time", "Post URL", "Impressions",
            "Engagements", "Notes", "Screenshot / File",
        ],
    },
    "HVU Insider — Instagram": {
        "color": "041E42",
        "columns": [
            "Drop Day", "Post Time", "Format", "Content Side",
            "Caption / Copy (Editable)", "Graphic/Asset Needed", "Graphic Status",
            "Asset Link (Google Drive)", "Hashtags", "Link in Bio Updated?",
            "Sponsor Tie-In", "Status", "Actual Post Time", "Post URL",
            "Impressions", "Engagements", "Notes", "Screenshot / File",
        ],
    },
    "dropt — Platform Content": {
        "color": "B8860B",
        "columns": [
            "Drop Day", "Content Type", "Content Side", "Title",
            "Description / Copy", "Assigned To", "Position Group", "Due Date",
            "Points Awarded", "Sponsor Tie-In", "Status", "Platform URL",
            "Asset Link (Google Drive)", "Notes", "Screenshot / File",
        ],
    },
    "dropt — Games": {
        "color": "B8860B",
        "columns": [
            "Drop Day", "Game Type", "Content Side", "Question #",
            "Question Text", "Answer / Options", "Correct Answer", "XP Awarded",
            "Position Group", "Sponsor Tie-In", "Upload Status", "Live Date", "Notes",
        ],
    },
    "Graphics Tracker": {
        "color": "666666",
        "columns": [
            "Drop Day", "Graphic Name", "Dimensions", "Type", "Content Side",
            "Position Group", "Assigned To", "Used By (Tabs)", "Drop Day",
            "Design Status", "Review Status", "Approved By",
            "Asset Link (Google Drive)", "Sponsor Tie-In", "Notes", "Final File",
        ],
    },
    "Video Tracker": {
        "color": "666666",
        "columns": [
            "Drop Day", "Video", "Dimensions", "Type", "Content Side",
            "Position Group", "Assigned To", "Used By (Tabs)", "Drop Day",
            "Design Status", "Review Status", "Approved By",
            "Asset Link (Google Drive)", "Sponsor Tie-In", "Notes", "Final File",
        ],
    },
    "Sponsorship Deliverables": {
        "color": "666666",
        "columns": [
            "Drop Day", "Sponsor", "Activation Type", "Content Side",
            "Position Group", "Description", "Platform Placement", "Deliverable",
            "Assigned To", "Due Date", "Status", "Asset Link (Google Drive)",
            "Revenue", "Notes",
        ],
    },
    "Swarm — 20 Accounts": {
        "color": "666666",
        "columns": [
            "Tier", "Account / Person", "Platform", "Position Group",
            "Target Post Time", "Suggested Copy", "Actual Copy (Editable)",
            "Status", "Post URL", "Estimated Reach", "Notes",
        ],
    },
    "HOW TO USE": {
        "color": "999999",
        "columns": ["Guide"],
        "static": True,
    },
}

# ── Column width defaults ─────────────────────────────────────────────────────
WIDTH_MAP = {
    "Copy": 55, "Caption": 55, "Description": 55, "Suggested Copy": 55,
    "Actual Copy": 55, "Question Text": 40, "Answer": 30, "Notes": 30,
    "Post URL": 30, "Asset Link": 30, "Platform URL": 30, "Hashtags": 25,
    "Status": 15, "Design Status": 15, "Review Status": 15, "Upload Status": 15,
    "Graphic Status": 15, "Post Status": 15, "Drop Day": 14, "Due Date": 14,
    "Post Time": 14, "Live Date": 14, "Actual Post Time": 14, "Date": 14,
}


def _get_col_width(col_name: str) -> float:
    """Determine column width based on column name keywords."""
    for keyword, width in WIDTH_MAP.items():
        if keyword.lower() in col_name.lower():
            return width
    return 16  # default


def _setup_header_row(ws, columns: list[str]) -> None:
    """Style the header row: navy fill, white bold, frozen."""
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = NAVY_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = _get_col_width(col_name)

    ws.freeze_panes = "A2"


def _add_status_validation(ws, columns: list[str]) -> None:
    """Add dropdown validation to status columns."""
    status_cols = [
        i for i, c in enumerate(columns, start=1)
        if "status" in c.lower() and "graphic" not in c.lower()
    ]
    if not status_cols:
        return

    dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Scheduled,Live,Skipped"',
        allow_blank=True,
    )
    dv.error = "Please select a valid status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)

    for col_idx in status_cols:
        col_letter = get_column_letter(col_idx)
        dv.add(f"{col_letter}2:{col_letter}500")


def _insert_separator_row(ws, position_group: str) -> int:
    """Insert a black separator row at the end of existing data. Returns the row number."""
    sep_row = ws.max_row + 1
    max_col = ws.max_column or 1

    for col in range(1, max_col + 1):
        cell = ws.cell(row=sep_row, column=col)
        cell.fill = BLACK_FILL

    label_cell = ws.cell(row=sep_row, column=1)
    label_cell.value = f"\u2014 {position_group.upper()} DROP \u2014"
    label_cell.font = WHITE_BOLD
    label_cell.alignment = CENTER

    if max_col > 1:
        ws.merge_cells(
            start_row=sep_row, start_column=1,
            end_row=sep_row, end_column=max_col,
        )

    return sep_row


def _populate_how_to_use(ws) -> None:
    """Write the static HOW TO USE reference tab."""
    ws.column_dimensions["A"].width = 100

    content = [
        "HVU INSIDER — Content Intel Tracker — HOW TO USE",
        "",
        "This workbook tracks every piece of content, graphic, video, game,",
        "sponsor activation, and swarm account across all position group drops.",
        "",
        "TAB GUIDE:",
        "  Mike V — Twitter: All tweets from Mike V's personal account",
        "  Mike V — Instagram: All IG posts from Mike V's personal account",
        "  HVU Insider — Twitter: All tweets from @HVU_Insider",
        "  HVU Insider — Instagram: All IG posts from @hvu_insider",
        "  dropt — Platform Content: Articles, videos, profiles uploaded to dropt",
        "  dropt — Games: Trivia, polls, pick'ems, challenges, keywords",
        "  Graphics Tracker: Every graphic asset with dimensions and status",
        "  Video Tracker: Every video asset with status and approvals",
        "  Sponsorship Deliverables: Sponsor activations per drop",
        "  Swarm — 20 Accounts: Swarm network posting tracker",
        "",
        "STATUS DROPDOWN VALUES:",
        "  Not Started | In Progress | Scheduled | Live | Skipped",
        "",
        "CONTENT SIDE VALUES:",
        "  Performance | Human | Engagement | Shared",
        "",
        "REUSE INSTRUCTIONS:",
        "  When adding a new position group, a BLACK SEPARATOR ROW marks",
        "  where the new data begins. Each tab accumulates data across all drops.",
        "  DO NOT create a new file — always append to this tracker.",
        "",
        "FILTERING:",
        "  Use the 'Position Group' column to filter by drop.",
        "  Use 'Content Side' to see Performance vs Human content.",
    ]

    for row_idx, line in enumerate(content, start=1):
        cell = ws.cell(row=row_idx, column=1, value=line)
        if row_idx == 1:
            cell.font = Font(name="Arial", bold=True, size=14, color="1B365D")
        elif line.endswith(":"):
            cell.font = Font(name="Arial", bold=True, size=11, color="1B365D")
        else:
            cell.font = Font(name="Arial", size=10)


def _generate_mike_v_twitter_rows(
    position: str, year: int, players: list[dict], drop_date: str,
) -> list[list[str]]:
    """Generate pre-filled rows for Mike V Twitter tab."""
    pos = position.upper()
    top_target = players[0]["name"] if players else "TBD"
    second_target = players[1]["name"] if len(players) > 1 else "TBD"

    return [
        [drop_date, "AM (Radio)", "Performance", f"Tease {pos} board — context narrative, set drop time",
         "None", "", "", "No", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "12 PM", "Performance", f"Context narrative post — why {pos} matters now",
         "None", "", "", "No", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "7 AM", "Performance", f'Redacted board (████), competition angle, "full intel at 10"',
         "Redacted Board Graphic", "Not Started", "", "No", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "6 PM", "Performance", f"Top target spotlight: {top_target} with stats",
         "Top Target Feature Card", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "3 PM", "Human", f"Personality/playlist/TikTok hook -> platform poll",
         "Personality Poll Graphic", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +1", "7 AM", "Performance", f"Hot take on {pos} room, poll CTA (+10 XP)",
         "None", "", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +1", "8 AM", "Performance", f"Second target spotlight: {second_target}, OV angle",
         "Second Target Card", "Not Started", "", "No", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +1", "3 PM", "Human", "Board thread tease -> forum push",
         "None", "", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +2", "6 PM", "Performance", f"Weekend debate prompt: Most realistic {pos} commit?",
         "None", "", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
    ]


def _generate_hvu_twitter_rows(
    position: str, year: int, players: list[dict], drop_date: str,
) -> list[list[str]]:
    """Generate pre-filled rows for HVU Insider Twitter tab."""
    pos = position.upper()
    top_target = players[0]["name"] if players else "TBD"
    second_target = players[1]["name"] if len(players) > 1 else "TBD"

    return [
        [drop_date, "10 AM", "Launch", "Performance", f"{year} {pos} Intel Drop is LIVE. Full scouting intel inside.",
         "HERO Graphic", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "12 PM", "Feature", "Performance", f"Target spotlight: {top_target}",
         "Top Target Feature Card", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "2 PM", "Feature", "Human", "Personality hook — who are these recruits off the field?",
         "Personality Poll Graphic", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [drop_date, "3 PM", "Narrative", "Performance", "Context narrative — the bigger picture",
         "Context Narrative Graphic", "Not Started", "", "No", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +1", "10 AM", "Feature", "Performance", f"Target #2: {second_target}",
         "Second Target Card", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +1", "2 PM", "Feature", "Human", "TikTok tease — recruit personality content",
         "None", "", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +1", "3 PM", "Engagement", "Engagement", "Sweepstakes reminder — earn XP, win prizes",
         "Sweepstakes Reminder Card", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +2", "10 AM", "Recap", "Performance", f"{pos} drop recap — what we learned",
         "Recap Graphic", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        [f"{drop_date} +2", "3 PM", "Engagement", "Engagement", "Keyword hunt — find it, earn +20 XP",
         "Keyword Hunt Graphic", "Not Started", "", "Yes", "", "Not Started", "", "", "", "", "", ""],
        ["Saturday", "12 PM", "Recap", "Performance", "Saturday recap — weekend roundup",
         "Recap Graphic", "Not Started", "", "No", "", "Not Started", "", "", "", "", "", ""],
    ]


def _generate_dropt_games_rows(
    position: str, year: int, players: list[dict], drop_date: str,
) -> list[list[str]]:
    """Generate pre-filled rows for dropt Games tab."""
    pos = position.upper()
    rows = []

    # Performance trivia
    for i in range(1, 7):
        rows.append([
            drop_date, "Trivia", "Performance", str(i),
            f"[{pos} trivia question {i} — to be written]", "[Options]", "[Answer]",
            "5", pos, "", "Not Started", drop_date, "",
        ])

    # Pick'Ems
    for i in range(1, 4):
        rows.append([
            drop_date, "Pick'Em", "Performance", str(i),
            f"[{pos} pick'em {i} — to be written]", "[Options]", "",
            "5", pos, "", "Not Started", drop_date, "",
        ])

    # Polls
    for i in range(1, 4):
        rows.append([
            drop_date, "Poll", "Performance", str(i),
            f"[{pos} poll {i} — to be written]", "[Options]", "",
            "10", pos, "", "Not Started", drop_date, "",
        ])

    # Challenge
    rows.append([
        drop_date, "Challenge", "Performance", "1",
        f"[{pos} challenge — to be written]", "", "",
        "1000 + 3 sweeps", pos, "", "Not Started", drop_date, "",
    ])

    # Keyword
    rows.append([
        drop_date, "Keyword", "Performance", "1",
        "[Keyword — to be embedded in video]", "", "[Keyword]",
        "20", pos, "", "Not Started", drop_date, "",
    ])

    # Human side
    rows.append([
        drop_date, "Personality Poll", "Human", "1",
        f"Which {pos} target's personality matches yours?", "[Player names]", "",
        "10", pos, "", "Not Started", drop_date, "",
    ])
    rows.append([
        drop_date, "Match Quiz", "Human", "1",
        f'Which {year} {pos} target are you?', "[Personality traits]", "",
        "10-40", pos, "", "Not Started", drop_date, "",
    ])
    rows.append([
        drop_date, "Lifestyle Poll", "Human", "1",
        f"Best pregame playlist from {pos} targets?", "[Player playlists]", "",
        "10", pos, "", "Not Started", drop_date, "",
    ])

    # Sponsor / Zero-Party
    rows.append([
        drop_date, "Data Capture", "Sponsor", "1",
        "[Sponsor data capture question 1]", "[Options]", "",
        "5", pos, "TBD", "Not Started", drop_date, "",
    ])
    rows.append([
        drop_date, "Data Capture", "Sponsor", "2",
        "[Sponsor data capture question 2]", "[Options]", "",
        "5", pos, "TBD", "Not Started", drop_date, "",
    ])

    return rows


def _generate_graphics_rows(
    position: str, year: int, players: list[dict], drop_date: str,
) -> list[list[str]]:
    """Generate pre-filled rows for Graphics Tracker tab."""
    pos = position.upper()
    graphics = [
        (f"HERO: Mike V + {pos} Board", "1080x1350 + 1600x900", "Hero", "Performance", "Day Before"),
        ("Redacted Board Graphic", "1080x1080", "Social", "Performance", "Day Before"),
        ("Landon x Mike V Split Design", "1080x1350 + 1080x1920", "Collab", "Performance", "Day Before"),
        ("IG Story Carousel (5 cards)", "1080x1920", "Story", "Performance", "Day Before"),
        ("Top Target Feature Card", "1080x1350", "Feature", "Performance", "Drop Day"),
        ("Context Narrative Graphic", "1080x1080", "Social", "Performance", "Drop Day"),
        ("Mike V Quote Card", "1080x1080 + 1080x1920", "Quote", "Performance", "Drop Day"),
        ("Personality Poll Graphic", "1080x1920", "Poll", "Human", "Drop Day"),
        ("Second Target Card", "1080x1350", "Feature", "Performance", "Drop Day"),
        ("Regional/Storyline Target Card", "1080x1350", "Feature", "Performance", "Drop Day"),
        ("Sweepstakes Reminder Card", "1080x1080 + 1080x1920", "Promo", "Engagement", "Drop +1"),
        ("Keyword Hunt Graphic", "1080x1920", "Promo", "Engagement", "Drop +1"),
        ("Quiz Promo Card", "1080x1080", "Promo", "Human", "Drop +1"),
        ("Recap Graphic", "1080x1080", "Recap", "Performance", "Drop +2"),
    ]

    rows = []
    for name, dims, gtype, side, due in graphics:
        rows.append([
            drop_date, name, dims, gtype, side, pos,
            "Ronny/Jonah", "", due, "Not Started", "Not Started",
            "Bryan", "", "", "", "",
        ])
    return rows


def _generate_swarm_rows(position: str) -> list[list[str]]:
    """Generate 20 swarm account placeholder rows."""
    pos = position.upper()
    rows = []
    for i in range(1, 21):
        tier = "Tier 1" if i <= 5 else ("Tier 2" if i <= 12 else "Tier 3")
        rows.append([
            tier, f"[Swarm Account {i}]", "X (Twitter)", pos,
            "Drop Day before 10 AM", "[Pre-written copy]", "",
            "Not Started", "", "", "",
        ])
    return rows


# Map tab names to their row generator functions
_ROW_GENERATORS = {
    "Mike V — Twitter": _generate_mike_v_twitter_rows,
    "HVU Insider — Twitter": _generate_hvu_twitter_rows,
    "dropt — Games": _generate_dropt_games_rows,
    "Graphics Tracker": _generate_graphics_rows,
}


def build_or_append_tracker(
    position: str,
    year: int,
    players: list[dict],
    drop_date: str,
    output_dir: Path,
) -> Path:
    """
    Build a new tracker or append to an existing one.

    If HVU_Intel_Content_Tracker.xlsx exists, inserts black separator rows
    on every tab (except HOW TO USE) and appends new position data below.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / "HVU_Intel_Content_Tracker.xlsx"
    pos = position.upper()
    existing = filepath.exists()

    if existing:
        wb = load_workbook(str(filepath))
    else:
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    for tab_name, tab_def in TAB_DEFINITIONS.items():
        columns = tab_def["columns"]
        is_static = tab_def.get("static", False)

        if existing and tab_name in wb.sheetnames:
            if is_static:
                continue  # Never modify HOW TO USE

            ws = wb[tab_name]

            # Insert black separator row
            _insert_separator_row(ws, pos)
            start_row = ws.max_row + 1

        else:
            ws = wb.create_sheet(title=tab_name)

            # Set tab color
            ws.sheet_properties.tabColor = tab_def["color"]

            if is_static:
                _populate_how_to_use(ws)
                continue

            # Set up header row
            _setup_header_row(ws, columns)
            _add_status_validation(ws, columns)
            start_row = 2

        # Generate and write data rows
        generator = _ROW_GENERATORS.get(tab_name)
        if generator:
            if tab_name == "Swarm — 20 Accounts":
                data_rows = _generate_swarm_rows(pos)
            elif tab_name == "Graphics Tracker":
                data_rows = generator(pos, year, players, drop_date)
            else:
                data_rows = generator(pos, year, players, drop_date)
        elif tab_name == "Swarm — 20 Accounts":
            data_rows = _generate_swarm_rows(pos)
        else:
            # Tabs without specific generators get placeholder rows
            data_rows = []

        for row_idx, row_data in enumerate(data_rows):
            actual_row = start_row + row_idx
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=actual_row, column=col_idx, value=value)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                # Alternating row colors
                if (actual_row % 2) == 0:
                    cell.fill = ALT_FILL

    wb.save(str(filepath))
    return filepath
